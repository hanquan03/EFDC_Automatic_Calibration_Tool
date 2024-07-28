# env3.8
"""
1) Creat a new ncfile which could contain EFDC simulation results 
   satisfying the behavioral parameter sets in daily scales.
2) Calculation spatial uncertainty of EFDC models for all grids in the study area.

"""


import os
import shutil
import math
import string
import netCDF4 as nc
import numpy as np
import pandas as pd
from datetime import datetime
import sys
from openpyxl import load_workbook, Workbook



#######################################################
# CREAT NEW NETCDF4 FILE
#======================================================

os.chdir(r"G:\AHydro_2016\code\Res\\")
hydro_res = nc.Dataset('500m_hydro_res_opt.nc', 'w', format='NETCDF4')
obj_fun = pd.read_excel(r"G:\AHydro_2016\code\Res\Obj_fun_res.xlsx")
opt_num = obj_fun[obj_fun['NSE'] > 0.5]
model_num = opt_num["ID"]


# Define dimesion
times = hydro_res.createDimension('time', size=None)
imax = hydro_res.createDimension('imax', size=58) 
jmax = hydro_res.createDimension('jmax', size=58) 
num = hydro_res.createDimension('num', size=len(model_num)) 


# Define variables
date = hydro_res.createVariable('Date', 'f4', dimensions='time')
number = hydro_res.createVariable('Number', 'i4', dimensions='num')
lon = hydro_res.createVariable('lon', 'f4', dimensions=("imax"))
lat = hydro_res.createVariable('lat', 'f4', dimensions=("jmax"))
test_n_age = hydro_res.createVariable('water_age', 'f4', dimensions=("num", 'time', 'imax', 'jmax'))
test_n_wl = hydro_res.createVariable('water_level', 'f4', dimensions=("num", 'time', 'imax', 'jmax'))
test_n_vx = hydro_res.createVariable('Vx', 'f4', dimensions=("num", 'time', 'imax', 'jmax'))
test_n_vy = hydro_res.createVariable('Vy', 'f4', dimensions=("num", 'time', 'imax', 'jmax'))


# Extract EFDC model outputs
res_wl = pro_wl = []
res_age = pro_age = []
res_vx = pro_vx = []
res_vy = pro_vy = []

for nn in range(len(model_num)):
    n = nn + 1
    nc_path = r"G:\\AHydro_2016\\code\\500m_Model\test" + str(model_num[nn]) + "\#output\DSI_EFDC.nc"
    nc_file = nc.Dataset(nc_path)

    data_wl = wl_data = []
    data_age = age_data = []
    data_vx = vx_data = []
    data_vy = vy_data = []


    for i in range(366):
        start = 24*i
        end = 24*(i+1)
        data_wl = nc_file.variables["WSEL"][start:end,:,:]
        data_wl[data_wl == -999] = np.NaN
        wl_data.append(np.nanmean(data_wl,axis=0))

        data_age = nc_file.variables["Dye"][start:end, :, :]
        data_age[data_age == -999] = np.NaN
        age_data.append(np.nanmean(data_age,axis=0))

        data_vx = nc_file.variables["Vx"][start:end, :, :]
        data_vx[data_vx == -999] = np.NaN
        vx_data.append(np.nanmean(data_vx,axis=0))

        data_vy = nc_file.variables["Vy"][start:end, :, :]
        data_vy[data_vy == -999] = np.NaN
        vy_data.append(np.nanmean(data_vy,axis=0))

    pro_wl.append(np.array(wl_data).reshape(366, 58, 58))
    pro_age.append(np.array(age_data).reshape(366, 58, 58))
    pro_vx.append(np.array(vx_data).reshape(366, 58, 58))
    pro_vy.append(np.array(vy_data).reshape(366, 58, 58))
    print ("Model number", len(pro_wl))

# Save simulation results
res_wl = np.array(pro_wl).reshape(len(model_num), 366, 58, 58)
res_age = np.array(pro_age).reshape(len(model_num), 366, 58, 58)
res_vx = np.array(pro_vx).reshape(len(model_num), 366, 58, 58)
res_vy = np.array(pro_vy).reshape(len(model_num), 366, 58, 58)
print ("1. Extract EFDC model outputs, Done")


# Add data to variables
lat[:] = np.linspace(38.7326,38.9847, num=58)
lon[:] = np.linspace(115.771,116.094, num=58)
number[:] = np.asarray(model_num)
time = []
date_range = pd.date_range(start="2015-10-1", end="2016-09-30", freq='D')
for i in range(int(float(366))):
    time.append(date_range[i].strftime('%Y-%m-%d'))
date = np.array(time).reshape(366,1)
test_n_wl[...] = pro_wl
test_n_age[...] = pro_age
test_n_vx[...] = pro_vx
test_n_vy[...] = pro_vy
print ("2. Write data, Done")


#Add global attributes
hydro_res.title = 'Create NetCDF file from model results by using netcdf4-python'
hydro_res.history = 'Created ' + datetime(2015, 10, 1).strftime('%Y-%m-%d')
hydro_res.Base_date = "2015-10-01"
hydro_res.End_date = "2016-09-30"
hydro_res.Project = "EFDC_model"
hydro_res.utm_zone = "UTM Zone 50 Northern Hemisphere"
hydro_res.Conventions = "CF-1.4"


#Add local attributes to variable
lon.standard_name = "longitude"
lon.description = 'longitude at grid cell center'
lon.units = 'degrees east'
lon.projection = "geographic"
lon.FillValue = "-999.0//float"

lat.standard_name = "latitude"
lat.description = 'latitude at grid cell center'
lat.units = 'degrees north'
lat.projection = "geographic"
lat.FillValue = "-999.0//float"

number.description = 'The model ID satisfying the behavioral parameter sets by automatic calibration method'
number.units = 'dimensionless'

test_n_wl.description = 'daily water level'
test_n_wl.units = 'm'
test_n_wl.coordinates = "lat lon"
test_n_wl.FillValue = "-999.0//float"

test_n_age.description = 'water age from basedate'
test_n_age.units = 'day'
test_n_age.coordinates = "lat lon"
test_n_age.FillValue = "-999.0//float"

test_n_vx.description = 'Eastward Water Velocity'
test_n_vx.units = 'm'
test_n_vx.coordinates = "lat lon"
test_n_vx.FillValue = "-999.0//float"

test_n_vy.description = 'Northward Water Velocity'
test_n_vy.units = 'm/s'
test_n_vy.coordinates = "lat lon"
test_n_vy.FillValue = "-999.0//float"


# Save result file
hydro_res.close()
print("3. Finish to create new NC file")


#######################################################
# CALCULATED SPATIAL UNCERTAINTY
#======================================================

def spatial_uncertainty(variable, output_path):

    """
    :param variable: Outputs of EFDC hydrodynamic model
    :param output_path: Path to store the results file
    :return: Spatial_UB.xlsx file
    """

    global Cell_ID, nc_file

    wl_hour = wl_model = []
    ss = np.empty(shape=(1, 3))

    # calculated spatial uncertainty for all grids
    # Extracting outputs of each grid in turn
    for n in range(len(Cell_ID)):
        i = Cell_ID[n][0] - 3
        j = Cell_ID[n][1] - 3
        wl_hour = nc_file.variables[variable][:, :, i, j]
        wl_model = np.nanmean(wl_hour, axis=1)
        wl_ID = nc_file.variables["Number"][:]

        df = pd.DataFrame({'ID': wl_ID, 'wl_model': wl_model})
        data = pd.merge(df, nse, on='ID', how='inner')

        if len(df) == len(nse):
            data = pd.merge(df, nse, on='ID', how='inner')
        else:
            data = pd.merge(df, nse, on='ID', how='right')

        data = pd.DataFrame(data)
        data.columns = ["ID", "wl", "NSE"]
        data["pro"] = data["NSE"].apply(lambda x: x / data["NSE"].sum())
        data = data.sort_values(by="wl", ascending=True)
        data["cum_pro"] = round(data["pro"].cumsum(), 2)

        sss = ()

        # Calculate the cumulative probability and obtain the quantile result
        index_5 = data[(data["cum_pro"] == min(data["cum_pro"], key=lambda x: abs(x - 0.05)))].iloc[0, 1]
        index_50 = data[(data["cum_pro"] == min(data["cum_pro"], key=lambda x: abs(x - 0.5)))].iloc[0, 1]
        index_95 = data[(data["cum_pro"] == min(data["cum_pro"], key=lambda x: abs(x - 0.95)))].iloc[0, 1]

        sss = np.array([[index_5, index_50, index_95]])
        ss = np.append(ss, sss, axis=0)
        print(variable, "cell ID is", n, i, j)

    ss = pd.DataFrame(ss).drop(index=[0])

    res = pd.concat([pd.DataFrame(Cell_ID), ss], axis=1)
    col_name = ('I', 'J', '5%quartile', '50%quartile', "95%quartile")
    res.columns = col_name
    res["UB_width"] = res["95%quartile"] - res["5%quartile"]


    # Loading an existing Excel file
    try:
        book = load_workbook(os.path.join(output_path, 'Spatial_UB.xlsx'))
    except FileNotFoundError:
        book = Workbook()

    writer = pd.ExcelWriter(os.path.join(output_path, 'Spatial_UB.xlsx'), engine='openpyxl')
    writer.book = book

    # Creating a new worksheet
    if variable in book.sheetnames:
        res.to_excel(writer, sheet_name=variable, startrow=book[variable].max_row, index=False)
    else:
        res.to_excel(writer, sheet_name=variable, index=False)

    writer.save()
    writer.close()


# Defining Global Variables
# read created NCfile, cell ID, and objective function result
Cell_ID = np.loadtxt(r"G:\Hydro_GLUE\GLUE_Python_Code\500m_cell_ID.txt")
nc_file = nc.Dataset(r"G:\AHydro_2016\code\Res\500m_hydro_res_opt.nc")

output_path = r"G:\AHydro_2016\code\Res\\"
nse = pd.read_excel(os.path.join(output_path, 'Obj_fun_res.xlsx'), usecols=["ID", "NSE"])
#nse = pd.read_excel(os.path.join(output_path, 'Obj_fun_res.xlsx'), usecols=[0, 3])



var = ("water_level", "water_age", "Vx", "Vy")
for var in var:
    ub = spatial_uncertainty(var, output_path)
    print(var, "Done")

print("4. EFDC Spatial uncertainty of all outputs has been completed")